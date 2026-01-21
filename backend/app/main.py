from __future__ import annotations

import asyncio
import json
import os
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from typing import Set

from fastapi import FastAPI, Request, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

from .models import Overview, EquipmentItem, AnalyticsResponse, Snapshot
from .simulator import Simulator

# Excel file path
EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", "Consumption.xlsx")


sim: Simulator | None = None
clients: Set[WebSocket] = set()
start_time = datetime.now(timezone.utc)


@asynccontextmanager
async def lifespan(app: FastAPI):
    global sim
    sim = Simulator()
    asyncio.create_task(broadcast_loop())
    yield
    # nothing to cleanup yet


app = FastAPI(
    title="Energy Monitor Demo",
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# static SPA (built frontend)
BASE_DIR = os.path.dirname(os.path.dirname(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "frontend_dist")
if os.path.isdir(STATIC_DIR):
    # Mount static files directory - this will serve assets, images, etc.
    app.mount("/assets", StaticFiles(directory=os.path.join(STATIC_DIR, "assets")), name="assets")
    # Serve other static files (like vite.svg) from root of frontend_dist
    static_files = StaticFiles(directory=STATIC_DIR)
    app.mount("/static", static_files, name="static")


@app.get("/")
async def index():
    index_path = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return {"message": "Frontend not built yet. Run frontend build."}


@app.get("/health")
async def health():
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}


@app.get("/api/overview", response_model=Overview)
async def api_overview():
    if not sim:
        raise RuntimeError("Simulator not initialized")
    snap = sim.generate_snapshot()
    uptime = int((datetime.now(timezone.utc) - start_time).total_seconds())
    return sim.build_overview(snap, uptime_seconds=uptime)


@app.get("/api/equipment", response_model=list[EquipmentItem])
async def api_equipment():
    if not sim:
        raise RuntimeError("Simulator not initialized")
    snap = sim.generate_snapshot()
    return sim.build_equipment(snap)


@app.get("/api/analytics", response_model=AnalyticsResponse)
async def api_analytics(hours: int = 24, resolution: int = 60):
    if not sim:
        raise RuntimeError("Simulator not initialized")
    return sim.build_analytics(hours=hours, resolution_minutes=resolution)


@app.get("/api/consumption-data")
async def get_consumption_data():
    """Read consumption data from Excel file and return current row data"""
    try:
        import openpyxl
        from datetime import datetime, timedelta
        
        if not os.path.exists(EXCEL_FILE):
            return {"error": "Consumption.xlsx not found"}
        
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb.active
        
        # Read path definitions from N3:R9 (columns N=14, O=15, P=16, Q=17, R=18)
        path_definitions = []
        for row_num in range(3, 10):  # Rows 3 to 9
            path_id = ws.cell(row=row_num, column=14).value  # Column N
            if path_id:
                path_definitions.append({
                    "path_id": str(path_id),
                    "from": str(ws.cell(row=row_num, column=15).value or "").lower(),  # Column O
                    "to": str(ws.cell(row=row_num, column=16).value or "").lower(),  # Column P
                    "color": str(ws.cell(row=row_num, column=17).value or ""),  # Column Q
                    "description": str(ws.cell(row=row_num, column=18).value or ""),  # Column R
                })
        
        # Find current time based on simulator (24 hours compressed to 2 minutes)
        # Calculate which 15-minute interval we're in
        if sim:
            elapsed = (datetime.now(timezone.utc) - start_time).total_seconds()
            # 24 hours = 86400 seconds, compressed to 120 seconds (2 minutes)
            # So 1 second real time = 720 seconds simulated time
            simulated_seconds = elapsed * 720
            # Find which 15-minute interval (0-95 intervals in 24 hours)
            interval = int(simulated_seconds / 900) % 96  # 900 seconds = 15 minutes
        else:
            interval = 0
        
        # Find data row (assuming data starts at row 2, with row 1 as header)
        # Adjust based on actual Excel structure
        data_start_row = 2
        current_row = data_start_row + interval
        
        # Read data row - adjust column indices based on actual Excel structure
        # Try to find columns by header names
        header_row = 1
        col_map = {}
        for col in range(1, 30):  # Check first 30 columns
            header = ws.cell(row=header_row, column=col).value
            if header:
                header_lower = str(header).lower()
                if "time" in header_lower:
                    col_map["time"] = col
                elif "building" in header_lower:
                    col_map["building"] = col
                elif "grid" in header_lower and "meter" not in header_lower:
                    col_map["grid"] = col
                elif "power" in header_lower:
                    col_map["power"] = col
                elif "solar" in header_lower:
                    col_map["solar"] = col
                elif "path 1" in header_lower or header_lower == "path1":
                    col_map["path1"] = col
                elif "path 2" in header_lower or header_lower == "path2":
                    col_map["path2"] = col
                elif "path 3" in header_lower or header_lower == "path3":
                    col_map["path3"] = col
        
        # Read values using column map
        time_value = ""
        if "time" in col_map:
            time_cell = ws.cell(row=current_row, column=col_map["time"])
            if time_cell.value:
                if isinstance(time_cell.value, datetime):
                    time_value = time_cell.value.strftime("%H:%M")
                else:
                    time_value = str(time_cell.value)
        
        building_kw = ws.cell(row=current_row, column=col_map.get("building", 2)).value or 0
        grid_kw = ws.cell(row=current_row, column=col_map.get("grid", 3)).value or 0
        power_kw = ws.cell(row=current_row, column=col_map.get("power", 4)).value or 0
        solar_kw = ws.cell(row=current_row, column=col_map.get("solar", 5)).value or 0
        
        # Read active paths
        active_paths = []
        if "path1" in col_map:
            path1 = ws.cell(row=current_row, column=col_map["path1"]).value
            if path1:
                active_paths.append(str(path1))
        if "path2" in col_map:
            path2 = ws.cell(row=current_row, column=col_map["path2"]).value
            if path2:
                active_paths.append(str(path2))
        if "path3" in col_map:
            path3 = ws.cell(row=current_row, column=col_map["path3"]).value
            if path3:
                active_paths.append(str(path3))
        
        # Read labels (if present in Excel)
        labels = {}
        # This will be populated based on actual Excel structure
        
        return {
            "time": time_value,
            "building_kw": float(building_kw) if building_kw else 0,
            "grid_kw": float(grid_kw) if grid_kw else 0,
            "power_kw": float(power_kw) if power_kw else 0,
            "solar_kw": float(solar_kw) if solar_kw else 0,
            "active_paths": active_paths,
            "path_definitions": path_definitions,
            "labels": labels,
        }
    except ImportError:
        return {"error": "openpyxl not installed"}
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}


@app.websocket("/ws/live")
async def ws_live(websocket: WebSocket):
    await websocket.accept()
    clients.add(websocket)
    try:
        # send one snapshot immediately
        if sim:
            snap = sim.generate_snapshot()
            await websocket.send_text(
                json.dumps(
                    {"type": "snapshot", "data": json.loads(Snapshot.model_validate(snap).model_dump_json())}
                )
            )

        while True:
            try:
                msg = await asyncio.wait_for(websocket.receive_text(), timeout=30.0)
                if msg == "ping":
                    await websocket.send_text("pong")
            except asyncio.TimeoutError:
                await websocket.send_text(
                    json.dumps({"type": "keepalive", "timestamp": datetime.now(timezone.utc).isoformat()})
                )
    except WebSocketDisconnect:
        pass
    except Exception:
        pass
    finally:
        clients.discard(websocket)


async def broadcast_loop():
    while True:
        await asyncio.sleep(2)
        if not sim or not clients:
            continue
        snap = sim.generate_snapshot()
        payload = json.dumps(
            {"type": "snapshot", "data": json.loads(Snapshot.model_validate(snap).model_dump_json())}
        )
        dead: Set[WebSocket] = set()
        for ws in clients:
            try:
                await ws.send_text(payload)
            except Exception:
                dead.add(ws)
        clients.difference_update(dead)


# Catch-all route for SPA routing - must be last
# This handles static files at root (like vite.svg) and SPA routes
@app.get("/{full_path:path}")
async def serve_spa(full_path: str):
    # Don't interfere with API routes or WebSocket (these are handled by their specific routes)
    if full_path.startswith("api/") or full_path.startswith("ws/"):
        from fastapi import HTTPException
        raise HTTPException(status_code=404)
    
    # Try to serve the requested file if it exists (e.g., vite.svg)
    file_path = os.path.join(STATIC_DIR, full_path)
    if os.path.exists(file_path) and os.path.isfile(file_path):
        return FileResponse(file_path)
    
    # Otherwise serve index.html for SPA routing
    index_path = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    
    return {"message": "Frontend not built yet. Run frontend build."}


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=True)
